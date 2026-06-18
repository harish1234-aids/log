import os
import json
import uuid
import io
import openpyxl
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_from_directory, send_file
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from sqlalchemy import func, extract
from datetime import datetime
from models import db, User, Student, Attendance, InternalMark

app = Flask(__name__)
# Configurations
app.config['SECRET_KEY'] = 'supersecretkey123'
base_dir = os.path.abspath(os.path.dirname(__file__))

# Database Configuration (Neon/Postgres for Vercel, SQLite for Local)
db_url = os.environ.get('DATABASE_URL')
if db_url:
    # SQLAlchemy 1.4+ requires 'postgresql://' instead of 'postgres://'
    if db_url.startswith("postgres://"):
        db_url = db_url.replace("postgres://", "postgresql://", 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = db_url
else:
    app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{os.path.join(base_dir, "attendance.db")}'

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# File upload config
UPLOAD_FOLDER = os.path.join(base_dir, 'static', 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB limit
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'gif', 'webp'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Initialize extensions
db.init_app(app)

login_manager = LoginManager()
login_manager.login_view = 'login'
login_manager.init_app(app)

with app.app_context():
    try:
        db.create_all()
        # Initialize principal admin dynamically safely
        principal = User.query.filter_by(role='principal').first()
        if not principal:
            principal = User(username='principal', role='principal')
            principal.set_password('principal123')
            db.session.add(principal)
            db.session.commit()
    except Exception as e:
        print(f"Database initialization deferred or failed: {e}")


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

@app.route('/')
def index():
    if current_user.is_authenticated:
        if current_user.role == 'principal':
            return redirect(url_for('manage_hod'))
        elif current_user.role == 'hod':
            return redirect(url_for('manage_staff'))
        elif current_user.role == 'staff':
            return redirect(url_for('admin_dashboard'))
        else:
            return redirect(url_for('student_home'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
        
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            login_user(user)
            return redirect(url_for('index'))
        else:
            flash('Invalid username or password', 'danger')
            
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# --- PRINCIPAL ONLY ROUTES ---
@app.route('/manage_hod', methods=['GET', 'POST'])
@login_required
def manage_hod():
    if current_user.role != 'principal':
        return redirect(url_for('index'))
        
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'add':
            username = request.form.get('username')
            password = request.form.get('password')
            dept = request.form.get('department')
            name = request.form.get('name')
            phone = request.form.get('phone')
            email = request.form.get('email')
            gender = request.form.get('gender')
            
            if User.query.filter_by(username=username).first():
                flash('Username already exists!', 'danger')
            else:
                hod = User(username=username, role='hod', department=dept, name=name, phone=phone, email=email, gender=gender)
                hod.set_password(password)
                db.session.add(hod)
                db.session.commit()
                flash(f'HOD for {dept} created successfully!', 'success')
                
        elif action == 'delete':
            user_id = request.form.get('user_id')
            user_to_delete = User.query.get(user_id)
            if user_to_delete and user_to_delete.role == 'hod':
                db.session.delete(user_to_delete)
                db.session.commit()
                flash('HOD deleted!', 'success')
                
        elif action == 'update':
            user_id = request.form.get('user_id')
            new_username = request.form.get('username')
            new_password = request.form.get('password')
            new_dept = request.form.get('department')
            name = request.form.get('name')
            phone = request.form.get('phone')
            email = request.form.get('email')
            gender = request.form.get('gender')

            hod_user = User.query.get(user_id)
            if hod_user and hod_user.role == 'hod':
                existing_user = User.query.filter_by(username=new_username).first()
                if existing_user and existing_user.id != hod_user.id:
                    flash('Username already taken.', 'danger')
                else:
                    hod_user.username = new_username
                    hod_user.department = new_dept
                    hod_user.name = name
                    hod_user.phone = phone
                    hod_user.email = email
                    hod_user.gender = gender
                    if new_password:
                        hod_user.set_password(new_password)
                    db.session.commit()
                    flash('HOD updated successfully!', 'success')

        elif action == 'import_hod':
            file = request.files.get('excel_file')
            if not file or not file.filename.endswith('.xlsx'):
                flash('Please upload a valid .xlsx file', 'danger')
            else:
                try:
                    wb = openpyxl.load_workbook(file)
                    ws = wb.active
                    added_count = 0
                    error_rows = []
                    
                    # Columns: Username, Password, Name, Department, Phone, Email, Gender
                    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                        if not row or all(cell is None for cell in row): continue
                        
                        username = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ''
                        password = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
                        name = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
                        dept = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''
                        phone = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ''
                        email = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ''
                        gender = str(row[6]).strip() if len(row) > 6 and row[6] is not None else ''
                        
                        if not username or not password or not dept:
                            error_rows.append(f'Row {row_num}: Missing Username, Password or Department')
                            continue
                        
                        if User.query.filter_by(username=username).first():
                            error_rows.append(f'Row {row_num}: Username "{username}" already exists')
                            continue
                            
                        hod = User(username=username, role='hod', department=dept, name=name, phone=phone, email=email, gender=gender)
                        hod.set_password(password)
                        db.session.add(hod)
                        added_count += 1
                    
                    if error_rows:
                        db.session.rollback()
                        flash('Import failed! Errors:\n' + '\n'.join(error_rows), 'danger')
                    else:
                        db.session.commit()
                        flash(f'Successfully imported {added_count} HODs!', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'Error: {str(e)}', 'danger')

    hod_members = User.query.filter_by(role='hod').all()
    departments = [d[0] for d in db.session.query(Student.department).distinct().all() if d[0]]
    if not departments: departments = ['CSE', 'IT', 'ECE', 'MECH']
    
    return render_template('manage_hod.html', hod_members=hod_members, departments=departments)

# --- HOD & PRINCIPAL SHARED ROUTES ---
@app.route('/manage_staff', methods=['GET', 'POST'])
@login_required
def manage_staff():
    if current_user.role not in ['principal', 'hod']:
        return redirect(url_for('index'))
    
    if request.method == 'POST' and current_user.role == 'principal':
        flash('Principals only have view access to staff management.', 'warning')
        return redirect(url_for('manage_staff'))
        
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'add':
            username = request.form.get('username')
            password = request.form.get('password')
            name = request.form.get('name')
            phone = request.form.get('phone')
            email = request.form.get('email')
            gender = request.form.get('gender')
            
            if User.query.filter_by(username=username).first():
                flash('Username already exists!', 'danger')
            else:
                staff = User(username=username, role='staff', 
                             department=current_user.department if current_user.role == 'hod' else None,
                             name=name, phone=phone, email=email, gender=gender)
                staff.set_password(password)
                db.session.add(staff)
                db.session.commit()
                flash('Staff account created successfully!', 'success')
                
        elif action == 'delete':
            user_id = request.form.get('user_id')
            user_to_delete = User.query.get(user_id)
            if user_to_delete and user_to_delete.role == 'staff':
                db.session.delete(user_to_delete)
                db.session.commit()
                flash('Staff deleted!', 'success')
                
        elif action == 'update':
            user_id = request.form.get('user_id')
            new_username = request.form.get('username')
            new_password = request.form.get('password') # optional
            name = request.form.get('name')
            phone = request.form.get('phone')
            email = request.form.get('email')
            gender = request.form.get('gender')

            staff_user = User.query.get(user_id)
            if staff_user and staff_user.role == 'staff':
                existing_user = User.query.filter_by(username=new_username).first()
                if existing_user and existing_user.id != staff_user.id:
                    flash('Username already taken.', 'danger')
                else:
                    staff_user.username = new_username
                    staff_user.name = name
                    staff_user.phone = phone
                    staff_user.email = email
                    staff_user.gender = gender
                    if new_password:
                        staff_user.set_password(new_password)
                    db.session.commit()
                    flash('Staff updated successfully!', 'success')

        elif action == 'import_staff':
            file = request.files.get('excel_file')
            target_dept = request.form.get('department')
            if current_user.role == 'hod':
                target_dept = current_user.department

            if not file or not file.filename.endswith('.xlsx'):
                flash('Please upload a valid .xlsx file', 'danger')
            elif not target_dept:
                flash('Please specify a target department', 'danger')
            else:
                try:
                    wb = openpyxl.load_workbook(file)
                    ws = wb.active
                    added_count = 0
                    error_rows = []
                    
                    # Columns: Username, Password, Name, Phone, Email, Gender
                    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                        if not row or all(cell is None for cell in row): continue
                        
                        username = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ''
                        password = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
                        name = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
                        phone = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''
                        email = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ''
                        gender = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ''
                        
                        if not username or not password:
                            error_rows.append(f'Row {row_num}: Missing Username or Password')
                            continue
                        
                        if User.query.filter_by(username=username).first():
                            error_rows.append(f'Row {row_num}: Username "{username}" already exists')
                            continue
                            
                        staff = User(username=username, role='staff', department=target_dept,
                                     name=name, phone=phone, email=email, gender=gender)
                        staff.set_password(password)
                        db.session.add(staff)
                        added_count += 1
                    
                    if error_rows:
                        db.session.rollback()
                        flash('Import failed! Errors:\n' + '\n'.join(error_rows), 'danger')
                    else:
                        db.session.commit()
                        flash(f'Successfully imported {added_count} staff members!', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'Error: {str(e)}', 'danger')

    query = User.query.filter_by(role='staff')
    if current_user.role == 'hod':
        from models import StaffAllocation
        staff_ids = db.session.query(StaffAllocation.staff_id).filter_by(department=current_user.department).distinct()
        query = query.filter((User.department == current_user.department) | (User.id.in_(staff_ids)))

    staff_members = query.all()
    departments = [d[0] for d in db.session.query(Student.department).distinct().all() if d[0]]
    if not departments: departments = ['CSE', 'IT', 'ECE', 'MECH']
    return render_template('manage_staff.html', staff_members=staff_members, departments=departments)

@app.route('/manage_allocations/<int:staff_id>', methods=['GET', 'POST'])
@login_required
def manage_allocations(staff_id):
    if current_user.role not in ['principal', 'hod']:
        return redirect(url_for('index'))
        
    if request.method == 'POST' and current_user.role == 'principal':
        flash('Principals only have view access to staff allocations.', 'warning')
        return redirect(url_for('manage_allocations', staff_id=staff_id))
        
    staff = User.query.get_or_404(staff_id)
    if staff.role != 'staff':
        return redirect(url_for('manage_staff'))
        
    if current_user.role == 'hod':
        # Check if HOD is allowed to manage this staff
        # For simplicity, if they are here, we allow it, but we'll restrict the department selection in the template
        pass
        
    from models import StaffAllocation
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'add':
            dept = request.form.get('department')
            year_val = request.form.get('year')
            subject = request.form.get('subject')
            
            if dept and year_val and subject:
                alloc = StaffAllocation(staff_id=staff.id, department=dept, year=int(year_val), subject=subject)
                db.session.add(alloc)
                db.session.commit()
                flash('Subject allocated successfully!', 'success')
                
        elif action == 'delete':
            alloc_id = request.form.get('allocation_id')
            alloc = StaffAllocation.query.get(alloc_id)
            if alloc and alloc.staff_id == staff.id:
                db.session.delete(alloc)
                db.session.commit()
                flash('Allocation removed.', 'success')
                
        return redirect(url_for('manage_allocations', staff_id=staff.id))
        
    allocations = StaffAllocation.query.filter_by(staff_id=staff.id).all()
    departments = [d[0] for d in Student.query.with_entities(Student.department).distinct().all() if d[0]]
    years = [y[0] for y in Student.query.with_entities(Student.year).distinct().order_by(Student.year).all() if y[0]]
    if not departments: departments = ['CSE', 'IT', 'ECE', 'MECH']
    if not years: years = [1, 2, 3, 4]
    
    return render_template('manage_allocations.html', staff=staff, allocations=allocations, departments=departments, years=years)


# --- STAFF & MASTER ADMIN ROUTES (CRUD Data) ---
@app.route('/download_template')
@login_required
def download_template():
    if current_user.role not in ['master_admin', 'staff']:
        return redirect(url_for('index'))
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"
    # Header
    headers = ["Name", "Department", "Year", "Username", "Password", "Phone", "Email", "Gender"]
    ws.append(headers)
    # Sample row
    ws.append(["Sample Student", "CSE", 1, "sample_user", "password123", "1234567890", "sample@example.com", "Male"])
    
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    return send_file(
        out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='student_import_template.xlsx'
    )

# --- STAFF & MASTER ADMIN ROUTES (CRUD Data) ---
@app.route('/manage_students', methods=['GET', 'POST'])
@login_required
def manage_students():
    if current_user.role not in ['principal', 'hod', 'staff']:
        return redirect(url_for('index'))
        
    if request.method == 'POST' and current_user.role == 'principal':
        flash('Principals only have view access to student management.', 'warning')
        return redirect(url_for('manage_students'))
    
    # HOD restriction
    if current_user.role == 'hod':
        user_dept = current_user.department
    else:
        user_dept = None
        
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'add':
            name = request.form.get('name')
            department = request.form.get('department')
            year = request.form.get('year')
            username = request.form.get('username')
            password = request.form.get('password')
            phone = request.form.get('phone')
            email = request.form.get('email')
            gender = request.form.get('gender')
            
            # Validate all required fields
            missing = []
            if not name: missing.append('Name')
            if not department: missing.append('Department')
            if not year: missing.append('Year')
            if not username: missing.append('Username')
            if not password: missing.append('Password')
            if not phone: missing.append('Phone')
            if not email: missing.append('Email')
            if not gender: missing.append('Gender')
            
            if missing:
                flash(f'Missing required fields: {", ".join(missing)}', 'danger')
            elif User.query.filter_by(username=username).first():
                flash('Username already exists!', 'danger')
            else:
                try:
                    student = Student(name=name, department=department, year=int(year), phone=phone, email=email, gender=gender)
                    db.session.add(student)
                    db.session.commit()
                    
                    user = User(username=username, role='student', department=department, year=int(year), student_id=student.id)
                    user.set_password(password)
                    db.session.add(user)
                    db.session.commit()
                    
                    flash('Student added successfully!', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'An error occurred: {str(e)}', 'danger')
                    
        elif action == 'delete':
            student_id = request.form.get('student_id')
            student = Student.query.get(student_id)
            if student:
                # Due to cascade delete, this will remove attendance records as well
                db.session.delete(student)
                db.session.commit()
                flash('Student and records deleted.', 'success')
                
        elif action == 'update':
            student_id = request.form.get('student_id')
            name = request.form.get('name')
            department = request.form.get('department')
            year = request.form.get('year')
            username = request.form.get('username')
            password = request.form.get('password')
            phone = request.form.get('phone')
            email = request.form.get('email')
            gender = request.form.get('gender')

            student = Student.query.get(student_id)
            if student:
                user = student.user_account
                if username and user.username != username:
                    if User.query.filter_by(username=username).first():
                        flash('Username already taken by another user!', 'danger')
                        return redirect(url_for('manage_students'))
                    user.username = username
                    
                student.name = name
                student.department = department
                student.year = int(year)
                student.phone = phone
                student.email = email
                student.gender = gender
                
                if user:
                    user.department = department
                    user.year = int(year)
                    if password:
                        user.set_password(password)

                db.session.commit()
                
                # Update attendance tags to stay visually consistent
                Attendance.query.filter_by(student_id=student.id).update({
                    'department': department,
                    'year': int(year)
                })
                db.session.commit()

                flash('Student updated successfully!', 'success')
            
        elif action == 'import_excel':
            file = request.files.get('excel_file')
            if not file or not file.filename.endswith('.xlsx'):
                flash('Please upload a valid .xlsx file', 'danger')
            else:
                try:
                    wb = openpyxl.load_workbook(file)
                    ws = wb.active
                    added_count = 0
                    error_rows = []
                    
                    field_names = ['Name', 'Department', 'Year', 'Username', 'Password', 'Phone', 'Email', 'Gender']
                    
                    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                        if not row or all(cell is None for cell in row):
                            continue
                        
                        # Extract all fields
                        raw_values = []
                        for i in range(8):
                            val = row[i] if len(row) > i and row[i] is not None else ''
                            raw_values.append(str(val).strip() if val != '' else '')
                        
                        name, department, year_str, username, password, phone, email, gender = raw_values
                        
                        # Validate all required fields (same as manual input)
                        missing = []
                        if not name: missing.append('Name')
                        if not department: missing.append('Department')
                        if not year_str: missing.append('Year')
                        if not username: missing.append('Username')
                        if not password: missing.append('Password')
                        if not phone: missing.append('Phone')
                        if not email: missing.append('Email')
                        if not gender: missing.append('Gender')
                        
                        if missing:
                            error_rows.append(f'Row {row_num}: Missing {", ".join(missing)}')
                            continue
                        
                        # Validate year is a number
                        try:
                            year_val = int(float(year_str))
                        except (ValueError, TypeError):
                            error_rows.append(f'Row {row_num}: Year must be a number')
                            continue
                        
                        # Check duplicate username
                        if User.query.filter_by(username=username).first():
                            error_rows.append(f'Row {row_num}: Username "{username}" already exists')
                            continue
                            
                        # HOD Department Isolation
                        if current_user.role == 'hod' and department != current_user.department:
                            error_rows.append(f'Row {row_num}: Department mismatch. You can only import students for {current_user.department}')
                            continue
                            
                        student = Student(name=name, department=department, year=year_val, phone=phone, email=email, gender=gender)
                        db.session.add(student)
                        db.session.flush()
                        
                        user = User(username=username, role='student', department=department, year=year_val, student_id=student.id)
                        user.set_password(password)
                        db.session.add(user)
                        added_count += 1
                    
                    if error_rows:
                        db.session.rollback()
                        error_msg = f'Import failed! Fix these errors in your Excel file:\n'
                        for err in error_rows:
                            error_msg += f'• {err}\n'
                        flash(error_msg, 'danger')
                    else:
                        db.session.commit()
                        flash(f'Successfully imported {added_count} students!', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'Error processing Excel file: {str(e)}', 'danger')

        elif action == 'promote':
            from_dept = request.form.get('from_dept')
            from_year = request.form.get('from_year')
            to_year = request.form.get('to_year')
            
            if from_dept and from_year and to_year and from_year.isdigit() and to_year.isdigit():
                from_year_int = int(from_year)
                to_year_int = int(to_year)
                
                students_to_promote = Student.query.filter_by(department=from_dept, year=from_year_int).all()
                count = 0
                for s in students_to_promote:
                    s.year = to_year_int
                    if s.user_account:
                        s.user_account.year = to_year_int
                    count += 1
                
                db.session.commit()
                # Update attendance tags to stay visually consistent
                Attendance.query.filter_by(department=from_dept, year=from_year_int).update({'year': to_year_int})
                db.session.commit()
                flash(f'Successfully promoted {count} students from {from_dept} Year {from_year_int} to Year {to_year_int}.', 'success')
            else:
                flash('Invalid promotion parameters.', 'danger')

        return redirect(url_for('manage_students'))

    # Retrieve Filter Arguments
    filter_dept = request.args.get('dept', 'All')
    filter_year = request.args.get('year', 'All')
    
    if current_user.role in ['hod', 'staff'] and current_user.department:
        filter_dept = current_user.department # Lock to user's department

    query = Student.query
    if filter_dept != 'All':
        query = query.filter(Student.department == filter_dept)
    if filter_year != 'All' and filter_year.isdigit():
        query = query.filter(Student.year == int(filter_year))
        
    students = query.all()
    
    if current_user.role in ['hod', 'staff'] and current_user.department:
        departments = [current_user.department]
    else:
        departments = [d[0] for d in Student.query.with_entities(Student.department).distinct().all() if d[0]]
    
    years = [y[0] for y in Student.query.with_entities(Student.year).distinct().order_by(Student.year).all() if y[0]]
    if not departments: departments = ['CSE', 'IT', 'ECE', 'MECH']
    if not years: years = [1, 2, 3, 4]

    return render_template('manage_students.html', students=students, 
                           departments=departments, years=years,
                           filter_dept=filter_dept, filter_year=filter_year)


@app.route('/admin', methods=['GET', 'POST'])
@login_required
def admin_dashboard():
    # Mark Attendance page
    if current_user.role not in ['principal', 'hod', 'staff']:
        return redirect(url_for('index'))
        
    if current_user.role == 'staff':
        from models import StaffAllocation
        allocs = StaffAllocation.query.filter_by(staff_id=current_user.id).all()
        departments = list(set([a.department for a in allocs]))
        years = list(set([a.year for a in allocs]))
    elif current_user.role == 'hod':
        departments = [current_user.department]
        years = [y[0] for y in db.session.query(Student.year).filter_by(department=current_user.department).distinct().all() if y[0]]
    else:
        # Principal
        departments = [d[0] for d in db.session.query(Student.department).distinct().all() if d[0]]
        years = [y[0] for y in db.session.query(Student.year).distinct().all() if y[0]]
    
    students = []
    selected_date = None
    selected_dept = None
    selected_year = None
    selected_period = None
    existing_attendance = {}
    
    if request.method == 'POST':
        action = request.form.get('action')
        selected_date = request.form.get('date')
        selected_dept = request.form.get('department')
        selected_year_str = request.form.get('year')
        selected_period_str = request.form.get('period')

        if current_user.role in ['hod', 'staff'] and current_user.department:
            selected_dept = current_user.department
        
        try:
            date_obj = datetime.strptime(selected_date, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            flash('Invalid date format', 'danger')
            return render_template('attend-admin.html', departments=departments, years=years, students=[], date=selected_date)
            
        if selected_year_str and selected_year_str.isdigit():
            selected_year = int(selected_year_str)
        if selected_period_str and selected_period_str.isdigit():
            selected_period = int(selected_period_str)
            
        if action == 'load':
            if selected_dept and selected_year and selected_period:
                students = Student.query.filter_by(department=selected_dept, year=selected_year).all()
                
                # Load existing attendance
                att_records = Attendance.query.filter_by(
                    department=selected_dept, year=selected_year, date=date_obj, period=selected_period
                ).all()
                
                for record in att_records:
                    existing_attendance[record.student_id] = record.status
                    
        elif action == 'submit':
            if current_user.role == 'principal':
                flash('Principals cannot mark attendance.', 'warning')
                return redirect(url_for('admin_dashboard'))
                
            if selected_dept and selected_year and selected_period:
                students = Student.query.filter_by(department=selected_dept, year=selected_year).all()
                for student in students:
                    status = request.form.get(f'attendance_{student.id}')
                    is_present = status == 'on' 

                    attendance_record = Attendance.query.filter_by(
                        student_id=student.id, date=date_obj, period=selected_period
                    ).first()
                    
                    if attendance_record:
                        attendance_record.status = is_present
                    else:
                        new_record = Attendance(
                            student_id=student.id,
                            date=date_obj,
                            period=selected_period,
                            status=is_present,
                            department=selected_dept,
                            year=selected_year
                        )
                        db.session.add(new_record)
                
                db.session.commit()
                flash(f'Attendance saved for Period {selected_period}!', 'success')
                
                # Reload existing
                att_records = Attendance.query.filter_by(
                    department=selected_dept, year=selected_year, date=date_obj, period=selected_period
                ).all()
                for record in att_records:
                    existing_attendance[record.student_id] = record.status

    from models import TimetableConfig, Feedback, SharedFile
    config = None
    if selected_dept and selected_year:
        config = TimetableConfig.query.filter_by(department=selected_dept, year=selected_year).first()
    timetable_blocks = []
    if config and config.layout_data:
        import json
        try:
            timetable_blocks = json.loads(config.layout_data)
        except:
            pass
            
    from sqlalchemy import or_
    if current_user.role == 'master_admin':
        feedbacks = Feedback.query.order_by(Feedback.timestamp.desc()).all()
    else:
        feedbacks = Feedback.query.filter(or_(Feedback.target_staff_id == None, Feedback.target_staff_id == current_user.id)).order_by(Feedback.timestamp.desc()).all()
    from models import News
    from sqlalchemy import or_
    if current_user.role == 'principal':
        news_posts = News.query.order_by(News.timestamp.desc()).all()
    elif current_user.role == 'hod':
        # See news for All, or specifically for HODs, or for their department
        news_posts = News.query.filter(
            or_(News.target_role == 'All', News.target_role == 'hod'),
            or_(News.target_dept == 'All', News.target_dept == current_user.department)
        ).order_by(News.timestamp.desc()).all()
    else:
        # Staff
        news_posts = News.query.filter(
            or_(News.target_role == 'All', News.target_role == 'staff'),
            or_(News.target_dept == 'All', News.target_dept == current_user.department)
        ).order_by(News.timestamp.desc()).all()
    shared_files = SharedFile.query.order_by(SharedFile.timestamp.desc()).all()

    all_students = Student.query.order_by(Student.name).all()
    return render_template('attend-admin.html', departments=departments, years=years, 
                           students=students, all_students=all_students, selected_date=selected_date,
                           selected_dept=selected_dept, selected_year=selected_year,
                           selected_period=selected_period, existing_attendance=existing_attendance,
                           timetable_blocks=timetable_blocks, feedbacks=feedbacks,
                           news_posts=news_posts, shared_files=shared_files)

@app.route('/delete_feedback/<int:fb_id>', methods=['POST'])
@login_required
def delete_feedback(fb_id):
    if current_user.role not in ['master_admin', 'staff']:
        return redirect(url_for('index'))
    from models import Feedback
    fb = Feedback.query.get(fb_id)
    if fb:
        db.session.delete(fb)
        db.session.commit()
        flash('Feedback deleted.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/post_news', methods=['POST'])
@login_required
def post_news():
    if current_user.role not in ['principal', 'hod', 'staff']:
        return redirect(url_for('index'))
    from models import News
    title = request.form.get('title')
    content = request.form.get('content')
    target_dept = request.form.get('target_dept') or 'All'
    target_year = request.form.get('target_year') or 'All'
    target_role = request.form.get('target_role') or 'All'
    
    if current_user.role in ['hod', 'staff'] and current_user.department:
        target_dept = current_user.department # Lock to user's department
        
    if title and content:
        n = News(author_id=current_user.id, title=title, content=content, 
                 target_dept=target_dept, target_year=target_year, target_role=target_role)
        db.session.add(n)
        db.session.commit()
        flash('News broadcast published successfully!', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/delete_news/<int:news_id>', methods=['POST'])
@login_required
def delete_news(news_id):
    if current_user.role not in ['principal', 'hod', 'staff']:
        return redirect(url_for('index'))
    from models import News
    n = News.query.get(news_id)
    if n and (n.author_id == current_user.id or current_user.role == 'principal'):
        db.session.delete(n)
        db.session.commit()
        flash('Announcement deleted.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/timetable_admin', methods=['GET', 'POST'])
@login_required
def timetable_admin():
    if current_user.role not in ['principal', 'hod']:
        return redirect(url_for('index'))

    if request.method == 'POST' and current_user.role == 'principal':
        flash('Principals only have view access to the timetable.', 'warning')
        return redirect(url_for('timetable_admin'))
        
    if current_user.role == 'hod':
        departments = [current_user.department]
        years = [y[0] for y in Student.query.with_entities(Student.year).filter_by(department=current_user.department).distinct().order_by(Student.year).all() if y[0]]
    else:
        departments = [d[0] for d in Student.query.with_entities(Student.department).distinct().all() if d[0]]
        years = [y[0] for y in Student.query.with_entities(Student.year).distinct().order_by(Student.year).all() if y[0]]
    if not departments: departments = ['CSE', 'IT', 'ECE', 'MECH']
    if not years: years = [1, 2, 3, 4]
    
    selected_dept = request.args.get('dept') or request.form.get('target_dept')
    if current_user.role in ['hod', 'staff'] and current_user.department:
        selected_dept = current_user.department
    
    selected_year_str = request.args.get('year') or request.form.get('year') or request.form.get('target_year')
    selected_year = int(selected_year_str) if selected_year_str and selected_year_str.isdigit() else None

    from models import TimetableConfig
    config = None
    
    if selected_dept and selected_year:
        config = TimetableConfig.query.filter_by(department=selected_dept, year=selected_year).first()
        if not config:
            config = TimetableConfig(department=selected_dept, year=selected_year)
            db.session.add(config)
            db.session.commit()
            
    if request.method == 'POST' and config:
        action = request.form.get('action')
        if action == 'generate':
            config.total_days = int(request.form.get('total_days') or 5)
            config.total_periods = int(request.form.get('total_periods') or 6)
            config.break_after = int(request.form.get('break_after') or 2)
            config.lunch_after = int(request.form.get('lunch_after') or 4)
            
            config.layout_data = None
            db.session.commit()
            flash('Timetable grid initialized!', 'success')
            return redirect(url_for('timetable_admin', dept=selected_dept, year=selected_year))
            
        elif action == 'save_names':
            layout_data = request.form.get('layout_data')
            if layout_data:
                config.layout_data = layout_data
                db.session.commit()
                flash('Timetable structures updated!', 'success')
            return redirect(url_for('timetable_admin', dept=selected_dept, year=selected_year))

        elif action == 'import_timetable':
            file = request.files.get('excel_file')
            if not file or not file.filename.endswith('.xlsx'):
                flash('Please upload a valid .xlsx file', 'danger')
            else:
                try:
                    import openpyxl, json
                    wb = openpyxl.load_workbook(file)
                    ws = wb.active
                    data_map = {} 
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row or all(c is None for c in row): continue
                        dept, year, day = row[0], row[1], row[2]
                        periods = [str(row[i]) if row[i] else '' for i in range(3, 11)]
                        key = (dept, year)
                        if key not in data_map: data_map[key] = []
                        data_map[key].append({'day': day, 'periods': periods})
                    for (dept, year), blocks in data_map.items():
                        c = TimetableConfig.query.filter_by(department=dept, year=int(year)).first()
                        if c:
                            c.layout_data = json.dumps({'columns': [], 'grid': blocks}) # Minimal format
                            c.total_days = len(blocks)
                        else:
                            c = TimetableConfig(department=dept, year=int(year), total_days=len(blocks), total_periods=8, layout_data=json.dumps({'columns': [], 'grid': blocks}))
                            db.session.add(c)
                    db.session.commit()
                    flash(f'Imported timetable for {len(data_map)} classes!', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'Error: {str(e)}', 'danger')
            return redirect(url_for('timetable_admin', dept=selected_dept, year=selected_year))

    return render_template('timetable_admin.html', config=config, departments=departments, years=years, selected_dept=selected_dept, selected_year=selected_year)

@app.route('/analytics')
@login_required
def analytics():
    if current_user.role not in ['principal', 'hod', 'staff']:
        return redirect(url_for('index'))
        
    date_str = request.args.get('date', datetime.today().strftime('%Y-%m-%d'))
    try:
        report_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        report_date = datetime.today().date()
        
    selected_dept = request.args.get('dept', 'All')
    if current_user.role in ['hod', 'staff'] and current_user.department:
        selected_dept = current_user.department
        
    # --- Data for Graph 1: Period-wise Attendance Graph of Total Students ---
    # Fetch total students marked present per period on the selected date
    query1 = db.session.query(
        Attendance.period, func.count(Attendance.id).label('total_present')
    ).filter(
        Attendance.date == report_date,
        Attendance.status == True
    )
    if selected_dept != 'All':
        query1 = query1.filter(Attendance.department == selected_dept)
    
    period_stats = query1.group_by(Attendance.period).order_by(Attendance.period).all()
    
    period_labels = [f"Period {p[0]}" for p in period_stats]
    period_data = [p[1] for p in period_stats]

    # --- Data for Graph 2: Monthly Attendance (Individual VS Overall) ---
    current_month = report_date.month
    current_year = report_date.year

    # Get overall monthly stats
    q_marked = Attendance.query.filter(
        extract('month', Attendance.date) == current_month,
        extract('year', Attendance.date) == current_year
    )
    q_present = Attendance.query.filter(
        extract('month', Attendance.date) == current_month,
        extract('year', Attendance.date) == current_year,
        Attendance.status == True
    )
    
    if selected_dept != 'All':
        q_marked = q_marked.filter(Attendance.department == selected_dept)
        q_present = q_present.filter(Attendance.department == selected_dept)

    total_monthly_marked = q_marked.count()
    total_monthly_present = q_present.count()

    overall_monthly_percentage = 0
    if total_monthly_marked > 0:
        overall_monthly_percentage = round((total_monthly_present / total_monthly_marked) * 100, 2)

    # Individual Student Table Data (Monthly)
    q_students = Student.query
    if selected_dept != 'All':
        q_students = q_students.filter(Student.department == selected_dept)
    students = q_students.all()
    
    monthly_student_stats = []
    
    for s in students:
        s_total = Attendance.query.filter_by(student_id=s.id).filter(
            extract('month', Attendance.date) == current_month,
            extract('year', Attendance.date) == current_year
        ).count()
        s_present = Attendance.query.filter_by(student_id=s.id, status=True).filter(
            extract('month', Attendance.date) == current_month,
            extract('year', Attendance.date) == current_year
        ).count()
        
        pct = round((s_present / s_total) * 100, 2) if s_total > 0 else 0
        monthly_student_stats.append({
            'name': s.name,
            'department': s.department,
            'total_periods': s_total,
            'present_periods': s_present,
            'percentage': pct
        })
        
    # Prepare top 5 for chart or just a subset
    top_students = sorted(monthly_student_stats, key=lambda x: x['percentage'], reverse=True)[:10]
    student_labels = [s['name'] for s in top_students]
    student_data = [s['percentage'] for s in top_students]

    if current_user.role == 'principal':
        departments = [d[0] for d in db.session.query(Student.department).distinct().all() if d[0]]
        years = [y[0] for y in db.session.query(Student.year).distinct().order_by(Student.year).all() if y[0]]
    elif current_user.role == 'hod':
        departments = [current_user.department]
        years = [y[0] for y in db.session.query(Student.year).filter_by(department=current_user.department).distinct().order_by(Student.year).all() if y[0]]
    else:
        # Staff
        from models import StaffAllocation
        allocs = StaffAllocation.query.filter_by(staff_id=current_user.id).all()
        departments = list(set([a.department for a in allocs]))
        years = sorted(list(set([a.year for a in allocs])))

    return render_template('analytics.html', 
                           report_date=report_date,
                           period_labels=period_labels,
                           period_data=period_data,
                           overall_monthly_percentage=overall_monthly_percentage,
                           monthly_student_stats=monthly_student_stats,
                           student_labels=student_labels,
                           student_data=student_data,
                           departments=departments,
                           years=years,
                           selected_dept=selected_dept)

# --- STUDENT ROUTES ---
@app.route('/home', methods=['GET', 'POST'])
@login_required
def student_home():
    if current_user.role != 'student':
        return redirect(url_for('index'))
        
    from models import Feedback, News
    if request.method == 'POST':
        message = request.form.get('message')
        target_staff_id = request.form.get('target_staff_id')
        if target_staff_id == '':
            target_staff_id = None
        if message:
            new_fb = Feedback(student_id=current_user.student_id, message=message, target_staff_id=target_staff_id)
            db.session.add(new_fb)
            db.session.commit()
            flash('Your feedback was safely submitted to the administration!', 'success')
            return redirect(url_for('student_home'))
            
    from sqlalchemy import or_
    latest_news = News.query.filter(
        or_(News.target_dept == 'All', News.target_dept == current_user.department),
        or_(News.target_year == 'All', News.target_year == str(current_user.year))
    ).order_by(News.timestamp.desc()).limit(10).all()

    from models import SharedFile
    student_files = SharedFile.query.filter(
        or_(SharedFile.target_dept == 'All', SharedFile.target_dept == current_user.department),
        or_(SharedFile.target_year == 'All', SharedFile.target_year == str(current_user.year))
    ).order_by(SharedFile.timestamp.desc()).all()
        
    current_month = datetime.today().month
    current_year = datetime.today().year
    
    total_periods = Attendance.query.filter_by(student_id=current_user.student_id).filter(
        extract('month', Attendance.date) == current_month,
        extract('year', Attendance.date) == current_year
    ).count()
    
    present_periods = Attendance.query.filter_by(student_id=current_user.student_id, status=True).filter(
        extract('month', Attendance.date) == current_month,
        extract('year', Attendance.date) == current_year
    ).count()
    
    monthly_percentage = 0
    if total_periods > 0:
        monthly_percentage = round((present_periods / total_periods) * 100, 2)
        
    from models import StaffAllocation
    allocs = StaffAllocation.query.filter_by(department=current_user.department, year=current_user.year).all()
    staff_ids = list(set([a.staff_id for a in allocs]))
    if staff_ids:
        staff_members = User.query.filter(User.id.in_(staff_ids)).all()
    else:
        staff_members = []
        
    return render_template('home.html', monthly_percentage=monthly_percentage,
                           latest_news=latest_news, student_files=student_files,
                           staff_members=staff_members)

@app.route('/attendance')
@login_required
def student_attendance_list():
    if current_user.role != 'student':
        return redirect(url_for('index'))
    
    # Dates where any attendance was recorded for this user
    query_dates = db.session.query(Attendance.date).filter_by(
        student_id=current_user.student_id
    ).distinct().order_by(Attendance.date.desc()).all()
    
    marked_dates = [d[0] for d in query_dates]
    return render_template('attendance.html', dates=marked_dates)

@app.route('/attendance/view')
@login_required
def student_attendance_view():
    if current_user.role != 'student':
        return redirect(url_for('index'))
        
    date_str = request.args.get('date')
    if not date_str:
        return redirect(url_for('student_attendance_list'))
        
    try:
        view_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return redirect(url_for('student_attendance_list'))
        
    # Get period-wise attendance for exactly this student on this day
    attendances_record = Attendance.query.filter_by(
        student_id=current_user.student_id,
        date=view_date
    ).order_by(Attendance.period).all()
        
    return render_template('attendance_view.html', attendances=attendances_record, date=view_date)

@app.route('/student_timetable')
@login_required
def student_timetable():
    if current_user.role != 'student':
        return redirect(url_for('index'))
        
    from models import TimetableConfig
    config = TimetableConfig.query.filter_by(department=current_user.department, year=current_user.year).first()
    timetable_blocks = []
    if config and config.layout_data:
        import json
        try:
            timetable_blocks = json.loads(config.layout_data)
        except:
            pass
            
    return render_template('timetable_student.html', timetable_blocks=timetable_blocks)

# --- INTERNAL MARKS SYSTEM ---

@app.route('/manage_marks', methods=['GET', 'POST'])
@login_required
def manage_marks():
    if current_user.role not in ['principal', 'hod', 'staff']:
        return redirect(url_for('index'))
        
    if request.method == 'POST' and current_user.role == 'principal':
        flash('Principals only have view access to marks management.', 'warning')
        return redirect(url_for('manage_marks'))
        
    from models import InternalMark
    
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            student_id = request.form.get('student_id')
            subject = request.form.get('subject')
            exam_type = request.form.get('exam_type')
            marks_obtained = request.form.get('marks_obtained')
            max_marks = request.form.get('max_marks', 100)
            
            im = InternalMark(student_id=student_id, author_id=current_user.id, subject=subject, exam_type=exam_type, marks_obtained=marks_obtained, max_marks=max_marks)
            db.session.add(im)
            db.session.commit()
            flash('Marks successfully recorded!', 'success')
        elif action == 'import_marks':
            file = request.files.get('excel_file')
            if not file or not file.filename.endswith('.xlsx'):
                flash('Please upload a valid .xlsx file', 'danger')
            else:
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file)
                    ws = wb.active
                    added_count = 0
                    error_rows = []
                    
                    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                        if not row or all(cell is None for cell in row): continue
                        
                        # Columns: Student Name, Dept, Year, Subject, Exam Type, Marks, Max Marks
                        raw_values = [str(row[i]).strip() if i < len(row) and row[i] is not None else '' for i in range(7)]
                        name, dept, year_str, subject, exam_type, marks_str, max_marks_str = raw_values
                        
                        if not all([name, dept, year_str, subject, exam_type, marks_str]):
                            error_rows.append(f'Row {row_num}: Missing required fields')
                            continue
                        
                        try:
                            year_val = int(float(year_str))
                            marks_val = float(marks_str)
                            max_marks_val = float(max_marks_str) if max_marks_str else 100.0
                        except:
                            error_rows.append(f'Row {row_num}: Invalid numeric values for Year/Marks/MaxMarks')
                            continue
                            
                        # Find student
                        student = Student.query.filter_by(name=name, department=dept, year=year_val).first()
                        if not student:
                            error_rows.append(f'Row {row_num}: Student "{name}" not found in {dept} Year {year_val}')
                            continue
                        
                        # Permissions
                        if current_user.role == 'hod' and dept != current_user.department:
                            error_rows.append(f'Row {row_num}: Dept mismatch. You can only import for {current_user.department}')
                            continue
                        
                        im = InternalMark(student_id=student.id, author_id=current_user.id, subject=subject, exam_type=exam_type, marks_obtained=marks_val, max_marks=max_marks_val)
                        db.session.add(im)
                        added_count += 1
                        
                    if error_rows:
                        db.session.rollback()
                        flash('Import failed! Errors:\n' + '\n'.join(error_rows[:10]), 'danger')
                    else:
                        db.session.commit()
                        flash(f'Successfully imported {added_count} mark records!', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'Error: {str(e)}', 'danger')

        elif action == 'delete':
            mark_id = request.form.get('mark_id')
            im = InternalMark.query.get(mark_id)
            if im and (current_user.role in ['principal', 'hod'] or im.author_id == current_user.id):
                db.session.delete(im)
                db.session.commit()
                flash('Mark record deleted.', 'success')

        return redirect(url_for('manage_marks'))
        
    # GET Filter Arguments
    filter_dept = request.args.get('dept', 'All')
    filter_year = request.args.get('year', 'All')
    
    if current_user.role in ['hod', 'staff'] and current_user.department:
        filter_dept = current_user.department # Lock to user's department
    
    q_students = Student.query
    if filter_dept != 'All':
        q_students = q_students.filter(Student.department == filter_dept)
    if filter_year != 'All' and filter_year.isdigit():
        q_students = q_students.filter(Student.year == int(filter_year))
    students = q_students.all()
    
    q_marks = InternalMark.query.join(Student)
    if filter_dept != 'All':
       q_marks = q_marks.filter(Student.department == filter_dept)
    if filter_year != 'All' and filter_year.isdigit():
       q_marks = q_marks.filter(Student.year == int(filter_year))
           
    marks = q_marks.order_by(InternalMark.timestamp.desc()).all()
    
    staff_subjects = []
    if current_user.role == 'staff':
        from models import StaffAllocation
        allocs = StaffAllocation.query.filter_by(staff_id=current_user.id).all()
        departments = list(set([a.department for a in allocs]))
        years = list(set([a.year for a in allocs]))
        staff_subjects = list(set([a.subject for a in allocs]))
    elif current_user.role == 'hod':
        departments = [current_user.department]
        years = [y[0] for y in Student.query.with_entities(Student.year).filter_by(department=current_user.department).distinct().order_by(Student.year).all() if y[0]]
        if not years: years = [1, 2, 3, 4]
    else:
        departments = [d[0] for d in Student.query.with_entities(Student.department).distinct().all() if d[0]]
        years = [y[0] for y in Student.query.with_entities(Student.year).distinct().order_by(Student.year).all() if y[0]]
        if not departments: departments = ['CSE', 'IT', 'ECE', 'MECH']
        if not years: years = [1, 2, 3, 4]
    
    return render_template('manage_marks.html', students=students, marks=marks, 
                           departments=departments, years=years, 
                           filter_dept=filter_dept, filter_year=filter_year,
                           staff_subjects=staff_subjects)

@app.route('/student_marks')
@login_required
def student_marks():
    if current_user.role != 'student':
        return redirect(url_for('index'))
        
    from models import InternalMark
    marks = InternalMark.query.filter_by(student_id=current_user.student_id).all()
    
    return render_template('student_marks.html', marks=marks)

# ── CHANGE CREDENTIALS ──

def _cred_redirect():
    """Safe redirect back to the caller's dashboard — works on mobile where Referer header may be absent."""
    if current_user.role == 'principal':
        return url_for('manage_hod')
    if current_user.role == 'hod':
        return url_for('manage_staff')
    return url_for('admin_dashboard')

@app.route('/change_credentials', methods=['POST'])
@login_required
def change_credentials():
    if current_user.role not in ['principal', 'hod', 'staff']:
        return redirect(url_for('index'))

    new_username     = request.form.get('new_username', '').strip()
    new_password     = request.form.get('new_password', '').strip()
    confirm_password = request.form.get('confirm_password', '').strip()

    if not new_username:
        flash('Username cannot be empty.', 'danger')
        return redirect(_cred_redirect())

    # Check username uniqueness (ignore current user's own record)
    existing = User.query.filter_by(username=new_username).first()
    if existing and existing.id != current_user.id:
        flash('That username is already taken. Please choose another.', 'danger')
        return redirect(_cred_redirect())

    # Check passwords match (only if a new password was supplied)
    if new_password:
        if new_password != confirm_password:
            flash('Passwords do not match. No changes were saved.', 'danger')
            return redirect(_cred_redirect())
        current_user.set_password(new_password)

    current_user.username = new_username
    try:
        db.session.commit()
        # REFRESH SESSION: Important for Flask-Login to recognize the updated user object
        from flask_login import login_user
        login_user(current_user._get_current_object())
        flash('Account credentials updated successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        error_msg = str(e)
        if 'readonly' in error_msg.lower():
            flash('Error: The database is read-only (common on Vercel). Changes cannot be saved.', 'danger')
        else:
            flash(f'An error occurred while saving: {error_msg}', 'danger')

    return redirect(_cred_redirect())


# ── FILE BROADCAST ROUTES ──

@app.route('/send_file', methods=['POST'])
@login_required
def send_file_broadcast():
    if current_user.role not in ['principal', 'hod', 'staff']:
        return redirect(url_for('index'))

    from models import SharedFile
    title = request.form.get('title', '').strip()
    target_dept = request.form.get('target_dept', 'All')
    if current_user.role in ['hod', 'staff'] and current_user.department:
        target_dept = current_user.department
    target_year = request.form.get('target_year', 'All')
    file = request.files.get('shared_file')

    if not title:
        flash('Please provide a title for the file.', 'danger')
        return redirect(url_for('admin_dashboard'))

    if not file or file.filename == '':
        flash('No file selected. Please choose a PDF or image.', 'danger')
        return redirect(url_for('admin_dashboard'))

    if not allowed_file(file.filename):
        flash('Invalid file type. Allowed: PDF, PNG, JPG, JPEG, GIF, WEBP.', 'danger')
        return redirect(url_for('admin_dashboard'))

    ext = file.filename.rsplit('.', 1)[1].lower()
    file_type = 'pdf' if ext == 'pdf' else 'image'
    unique_name = f"{uuid.uuid4().hex}_{file.filename}"
    save_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_name)
    file.save(save_path)

    sf = SharedFile(
        author_id=current_user.id,
        title=title,
        filename=unique_name,
        original_name=file.filename,
        file_type=file_type,
        target_dept=target_dept,
        target_year=target_year
    )
    db.session.add(sf)
    db.session.commit()
    flash(f'File "{title}" sent successfully!', 'success')
    return redirect(url_for('admin_dashboard'))


@app.route('/delete_shared_file/<int:file_id>', methods=['POST'])
@login_required
def delete_shared_file(file_id):
    if current_user.role not in ['principal', 'hod', 'staff']:
        return redirect(url_for('index'))

    from models import SharedFile
    sf = SharedFile.query.get(file_id)
    if sf and (current_user.role in ['principal', 'hod'] or sf.author_id == current_user.id):
        # Remove file from disk
        disk_path = os.path.join(app.config['UPLOAD_FOLDER'], sf.filename)
        if os.path.exists(disk_path):
            os.remove(disk_path)
        db.session.delete(sf)
        db.session.commit()
        flash('File deleted successfully.', 'success')
    else:
        flash('You do not have permission to delete this file.', 'danger')
    return redirect(url_for('admin_dashboard'))


@app.route('/uploads/<path:filename>')
@login_required
def serve_upload(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)



@app.route('/export_attendance')
@login_required
def export_attendance():
    if current_user.role not in ['principal', 'hod', 'staff']:
        return redirect(url_for('index'))
    
    dept = request.args.get('dept')
    year = request.args.get('year')
    student_id = request.args.get('student_id')
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    query = Attendance.query.join(Student)
    
    if current_user.role == 'hod' and (not dept or dept == 'All'):
        dept = current_user.department
        
    if dept and dept != 'All':
        query = query.filter(Attendance.department == dept)
    if year and year != 'All':
        query = query.filter(Attendance.year == int(year))
    if student_id and student_id != 'All':
        query = query.filter(Attendance.student_id == int(student_id))
    if start_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        query = query.filter(Attendance.date >= start_date)
    if end_date_str:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        query = query.filter(Attendance.date <= end_date)
        
    records = query.order_by(Attendance.date.desc(), Attendance.period.asc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    headers = ["Date", "Student Name", "Roll No", "Department", "Year", "Period", "Status"]
    ws.append(headers)
    
    for r in records:
        ws.append([
            r.date.strftime('%Y-%m-%d'),
            r.student.name,
            r.student_id,
            r.department,
            r.year,
            r.period,
            "Present" if r.status else "Absent"
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"attendance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)

@app.route('/export_marks')
@login_required
def export_marks():
    if current_user.role not in ['principal', 'hod', 'staff']:
        return redirect(url_for('index'))
    
    dept = request.args.get('dept')
    year = request.args.get('year')
    student_id = request.args.get('student_id')
    subject = request.args.get('subject')
    exam_type = request.args.get('exam_type')
    
    query = InternalMark.query.join(Student)
    
    if dept and dept != 'All':
        query = query.filter(Student.department == dept)
    if year and year != 'All':
        query = query.filter(Student.year == int(year))
    if student_id and student_id != 'All':
        query = query.filter(InternalMark.student_id == int(student_id))
    if subject and subject != 'All':
        query = query.filter(InternalMark.subject == subject)
    if exam_type and exam_type != 'All':
        query = query.filter(InternalMark.exam_type == exam_type)
        
    records = query.order_by(InternalMark.timestamp.desc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Internal Marks Report"
    
    headers = ["Student Name", "Roll No", "Department", "Year", "Subject", "Exam Type", "Marks Obtained", "Max Marks", "Percentage", "Date"]
    ws.append(headers)
    
    for r in records:
        percentage = (r.marks_obtained / r.max_marks * 100) if r.max_marks > 0 else 0
        ws.append([
            r.student.name,
            r.student_id,
            r.student.department,
            r.student.year,
            r.subject,
            r.exam_type,
            r.marks_obtained,
            r.max_marks,
            f"{percentage:.2f}%",
            r.timestamp.strftime('%Y-%m-%d %H:%M')
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"marks_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)


@app.route('/api/analytics')
@login_required
def api_analytics():
    role = current_user.role
    level = request.args.get('level', '1')
    target_dept = request.args.get('dept')
    target_year = request.args.get('year')
    target_subject = request.args.get('subject')
    search_query = request.args.get('search', '').strip()
    target_exam = request.args.get('exam', 'All')

    from models import Student, InternalMark, Attendance, StaffAllocation
    from sqlalchemy import func

    data = {
        'labels': [],
        'marks': [],
        'attendance': [],
        'students': [] # For level 3
    }

    def get_avg_marks(query):
        if target_exam != 'All':
            query = query.filter(InternalMark.exam_type == target_exam)
        return query.scalar() or 0

    if search_query:
        # Global student search across the authorized scope
        students_query = Student.query
        if role == 'hod':
            students_query = students_query.filter_by(department=current_user.department)
        elif role == 'staff':
            from models import StaffAllocation
            allocs = StaffAllocation.query.filter_by(staff_id=current_user.id).all()
            depts = [a.department for a in allocs]
            students_query = students_query.filter(Student.department.in_(depts))
        
        students = students_query.filter(Student.name.ilike(f"%{search_query}%")).all()
        for s in students:
            s_data = {
                'name': s.name,
                'marks_data': [],
                'attendance_pct': 0
            }
            marks_query = InternalMark.query.filter_by(student_id=s.id)
            if target_exam != 'All':
                marks_query = marks_query.filter_by(exam_type=target_exam)
            
            marks = marks_query.all()
            subj_marks = {}
            for m in marks:
                if m.subject not in subj_marks: subj_marks[m.subject] = []
                subj_marks[m.subject].append(m.marks_obtained / m.max_marks * 100)
            for subj, vals in subj_marks.items():
                s_data['marks_data'].append({'subject': subj, 'avg': round(sum(vals)/len(vals), 2)})
            
            att_total = Attendance.query.filter_by(student_id=s.id).count()
            att_present = Attendance.query.filter_by(student_id=s.id, status=True).count()
            s_data['attendance_pct'] = round((att_present / att_total * 100), 2) if att_total > 0 else 0
            data['students'].append(s_data)
        return jsonify(data)

    if role == 'principal':
        if level == '1':
            # Dept-wise averages
            depts = db.session.query(Student.department).distinct().all()
            for (d,) in depts:
                data['labels'].append(d)
                # Avg Marks
                m_q = db.session.query(func.avg(InternalMark.marks_obtained / InternalMark.max_marks * 100)).join(Student).filter(Student.department == d)
                data['marks'].append(round(get_avg_marks(m_q), 2))
                # Avg Attendance
                a_avg = db.session.query(func.avg(Attendance.status.cast(db.Float) * 100)).filter(Attendance.department == d).scalar() or 0
                data['attendance'].append(round(a_avg, 2))
        
        elif level == '2' and target_dept:
            # Year-wise averages in Dept
            years = db.session.query(Student.year).filter_by(department=target_dept).distinct().all()
            for (y,) in years:
                data['labels'].append(f"Year {y}")
                m_q = db.session.query(func.avg(InternalMark.marks_obtained / InternalMark.max_marks * 100)).join(Student).filter(Student.department == target_dept, Student.year == y)
                data['marks'].append(round(get_avg_marks(m_q), 2))
                a_avg = db.session.query(func.avg(Attendance.status.cast(db.Float) * 100)).filter(Attendance.department == target_dept, Attendance.year == y).scalar() or 0
                data['attendance'].append(round(a_avg, 2))

        elif level == '3' and target_dept and target_year:
            # Each student in that class
            students = Student.query.filter_by(department=target_dept, year=int(target_year)).all()
            for s in students:
                s_data = {
                    'name': s.name,
                    'marks_data': [], # Subject-wise for this student
                    'attendance_pct': 0
                }
                # Subject breakdown
                marks_q = InternalMark.query.filter_by(student_id=s.id)
                if target_exam != 'All':
                    marks_q = marks_q.filter_by(exam_type=target_exam)
                
                marks = marks_q.all()
                subj_marks = {}
                for m in marks:
                    if m.subject not in subj_marks: subj_marks[m.subject] = []
                    subj_marks[m.subject].append(m.marks_obtained / m.max_marks * 100)
                
                for subj, vals in subj_marks.items():
                    s_data['marks_data'].append({'subject': subj, 'avg': round(sum(vals)/len(vals), 2)})
                
                # Attendance
                att_total = Attendance.query.filter_by(student_id=s.id).count()
                att_present = Attendance.query.filter_by(student_id=s.id, status=True).count()
                s_data['attendance_pct'] = round((att_present / att_total * 100), 2) if att_total > 0 else 0
                data['students'].append(s_data)

    elif role == 'hod':
        dept = current_user.department
        if level == '1':
            # Subject-wise averages in Dept
            q = db.session.query(InternalMark.subject).join(Student).filter(Student.department == dept)
            if target_year:
                q = q.filter(Student.year == int(target_year))
            subjects = q.distinct().all()
            
            for (s,) in subjects:
                data['labels'].append(s)
                m_q = db.session.query(func.avg(InternalMark.marks_obtained / InternalMark.max_marks * 100)).join(Student).filter(
                    InternalMark.subject == s,
                    Student.department == dept
                )
                if target_year:
                    m_q = m_q.filter(Student.year == int(target_year))
                    
                data['marks'].append(round(get_avg_marks(m_q), 2))
                
                # Attendance average for this subject/year
                a_q = db.session.query(func.avg(Attendance.status.cast(db.Float) * 100)).filter(Attendance.department == dept)
                if target_year:
                    a_q = a_q.filter(Attendance.year == int(target_year))
                
                data['attendance'].append(round(a_q.scalar() or 0, 2)) 

        elif level == '2' and target_subject:
            # Individual students for this subject
            # Here we might want to show actual marks for the subject, or student avg
            marks_query = InternalMark.query.filter_by(subject=target_subject).join(Student).filter(Student.department == dept)
            if target_exam != 'All':
                marks_query = marks_query.filter(InternalMark.exam_type == target_exam)
            
            records = marks_query.all()
            for r in records:
                data['labels'].append(r.student.name)
                data['marks'].append(round(r.marks_obtained / r.max_marks * 100, 2))

    elif role == 'staff':
        # Staff handling subjects
        q_alloc = StaffAllocation.query.filter_by(staff_id=current_user.id)
        if target_year:
            q_alloc = q_alloc.filter_by(year=int(target_year))
        allocs = q_alloc.all()
        
        if level == '1':
            for a in allocs:
                label = f"{a.subject} ({a.department} Y{a.year})"
                data['labels'].append(label)
                m_q = db.session.query(func.avg(InternalMark.marks_obtained / InternalMark.max_marks * 100)).join(Student).filter(
                    InternalMark.subject == a.subject, Student.department == a.department, Student.year == a.year
                )
                data['marks'].append(round(get_avg_marks(m_q), 2))
                
                a_avg = db.session.query(func.avg(Attendance.status.cast(db.Float) * 100)).filter(
                    Attendance.department == a.department, Attendance.year == a.year
                ).scalar() or 0
                data['attendance'].append(round(a_avg, 2))

        elif level == '2' and target_dept and target_year:
            students = Student.query.filter_by(department=target_dept, year=int(target_year)).all()
            for s in students:
                data['labels'].append(s.name)
                m_q = db.session.query(func.avg(InternalMark.marks_obtained / InternalMark.max_marks * 100)).filter(InternalMark.student_id == s.id)
                data['marks'].append(round(get_avg_marks(m_q), 2))
                
    return jsonify(data)

if __name__ == '__main__':
    app.run(debug=True, port=5000)
